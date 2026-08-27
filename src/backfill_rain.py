"""Backfill historical rain data from a DHMZ spreadsheet export (mrse.xlsx).

Sheet layout: one sheet per city, repeating year blocks. Each block has a header
row (year in col A, months I-XII in cols B-M), then 31 day rows, then summary
rows. A blank day cell means 0 mm -- verified against the sheet's own STD row.
A month whose ZBROJ (sum) cell is '-' has no data at all and is skipped.
"""
import calendar
import os
import sys

import openpyxl

from src import db_handler
from src.config import BASE_DIR

# Sheet name -> (city, station). The sheet titles name the stations:
# "Zagreb - Pleso aerodrom" is the same station DHMZ publishes as
# "Zagreb-aerodrom", so history merges into that series. Split-Marjan is a
# distinct station from Split-aerodrom, so it stays its own series.
SHEETS = {
    'zagreb': ('Zagreb', 'Zagreb-aerodrom'),
    'split': ('Split', 'Split-Marjan'),
}

DAYS_PER_BLOCK = 31
ZBROJ_OFFSET = 33


def _year_header_rows(ws):
    return [r for r in range(1, ws.max_row + 1)
            if isinstance(ws.cell(r, 1).value, int) and ws.cell(r, 1).value > 1900]


def parse_sheet(ws):
    """Yield (iso_date, mm) for every day with data."""
    for header in _year_header_rows(ws):
        year = ws.cell(header, 1).value
        for col in range(2, 14):
            month = col - 1
            if ws.cell(header + ZBROJ_OFFSET, col).value == '-':
                continue  # month has no data
            for day in range(1, calendar.monthrange(year, month)[1] + 1):
                value = ws.cell(header + day, col).value
                if value is None:
                    mm = 0.0
                elif isinstance(value, (int, float)):
                    mm = float(value)
                else:
                    continue  # '-' or '*': day not measured
                yield f"{year}-{month:02d}-{day:02d}", mm


def main(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    conn = db_handler.setup_db()
    try:
        for sheet, (city, station) in SHEETS.items():
            count = 0
            for date, mm in parse_sheet(wb[sheet]):
                db_handler.insert_into_rain_db(conn, station, city, mm, date)
                count += 1
            print(f"[backfill] {station}: {count} days")
    finally:
        conn.close()


if __name__ == "__main__":
    main(sys.argv[1] if len(sys.argv) > 1 else os.path.join(BASE_DIR, "mrse.xlsx"))
